import pandas as pd
#Text files generator code is present in the below module
import Text_File_Generator
from Utils import Word_Utils
from openpyxl import load_workbook
from nltk.tokenize import sent_tokenize, word_tokenize
from Trie_Base import STOP_WORDS_TRIE, POSITIVE_WORDS_TRIE, NEGATIVE_WORDS_TRIE

#If file is not found. The output will be populated using the list below
File_Not_Found_Data_Frame = pd.DataFrame([["File Not Found"]*13])
Personal_Pronouns = ["i", "we", "my", "ours", "us"]

#All the text files are created inside this folder
Folder_Path = 'Text_Files/'
Extension = ".txt"
Output_File_Path = 'Output Data Structure.xlsx'
Log_File = Text_File_Generator.Log_File

#Reading the output data structure
Output_File = pd.read_excel(Output_File_Path)

#Iterating over each row of the Output file
for index, row in Output_File.iterrows():
    
    #Initializing the output variables required
    CLEAN_WORD_COUNT = 0
    POSITIVE_SCORE = 0
    NEGATIVE_SCORE = 0
    COMPLEX_WORD_COUNT = 0
    TOTAL_SYLLABLE_COUNT = 0 
    TOTAL_CHARACTER_COUNT = 0
    PERSONAL_PRONOUN_COUNT = 0
    
    #Cleaning the URL because File name can't contain same URL as its name
    simplified_URL = Word_Utils.URL_Cleaner(row['URL'])
    #Formatting the file path which is given to access the file
    File_Path = f"{Folder_Path}{simplified_URL}_{row['URL_ID']}{Extension}"
    
    try:
        #Opening the file
        File = open(File_Path, "r", encoding="utf8")
        
    except FileNotFoundError:
        
        #if the file is not found then populating the output as File Not Found
        print("File not found: " + File_Path)
        Log_File.log_event("File not found: " + File_Path)
        
        #Creating a writer for the output file
        with pd.ExcelWriter(Output_File_Path, engine='openpyxl', 
                            mode='a') as writer:
            book = load_workbook(Output_File_Path)
            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            File_Not_Found_Data_Frame.to_excel(writer, sheet_name='Sheet1', 
                                               header=None, index=False, 
                                               startcol=2,startrow=index+1)
            writer.save()
            File.close()
            Log_File.log_event(Text_File_Generator.Separator)
            
        #If the File is not found then moving on to the next file
        continue
    
    #If the file is present then reading the file
    File_Text = File.read()
    
    #Tokenizing the file into the sentences
    Sentences = sent_tokenize(File_Text)
    Sentence_Count = len(Sentences)
    
    #Tokenizing the file into words
    Word_List = word_tokenize(File_Text)
    Word_Count = len(Word_List)
    
    #For each word found performing the analytics    
    for word in Word_List:
        
        # Changing the word into lower case because the Tries are created by 
        # lowering the case.
        word = word.lower()
        
        #Performing the analysis only if the word is not a STOP WORD.
        if not STOP_WORDS_TRIE.search(word):
            
            #If word is not a stop word consider it as a clean word.
            CLEAN_WORD_COUNT += 1
            #Count the syllables of the word to know if it is a complex word.
            SYLLABLE_COUNT = Word_Utils.Syllable_Count(word)
            
            if SYLLABLE_COUNT > 1:
                #Word having more than one syllable is a complex word
                COMPLEX_WORD_COUNT = COMPLEX_WORD_COUNT+1
                
            #Checking if word isa Personal Pronoun
            if word in Personal_Pronouns:
                PERSONAL_PRONOUN_COUNT += 1 
                
            #Checking if word is a positive word
            if POSITIVE_WORDS_TRIE.search(word):
                POSITIVE_SCORE = POSITIVE_SCORE + 1
                
            #Checking if word is a negative word
            elif NEGATIVE_WORDS_TRIE.search(word):
                NEGATIVE_SCORE = NEGATIVE_SCORE + 1  
            
            #Maintaining the total syllable count across the file
            TOTAL_SYLLABLE_COUNT += SYLLABLE_COUNT
            #Maintaining the total character count across the file
            TOTAL_CHARACTER_COUNT += len(word)
    
    #Calculating the required output variables
    SCORE_SUM = POSITIVE_SCORE + NEGATIVE_SCORE
    SCORE_DIFF = POSITIVE_SCORE - NEGATIVE_SCORE
    POLARITY_SCORE =  SCORE_DIFF/(SCORE_SUM + 0.000001)
    SUBJECTIVITY_SCORE = SCORE_SUM/(CLEAN_WORD_COUNT + 0.000001)
    AVERAGE_SENTENCE_LENGTH = Word_Count/Sentence_Count
    PERCENTAGE_OF_COMPLEX_WORDS = (COMPLEX_WORD_COUNT/Word_Count)*100
    FOG_INDEX = 0.4 * (AVERAGE_SENTENCE_LENGTH + PERCENTAGE_OF_COMPLEX_WORDS)
    CLEAN_WORD_COUNT = Word_Utils.Word_Cleaner(Word_List)
    SYLLABLE_COUNT_PER_WORD = TOTAL_SYLLABLE_COUNT/Word_Count
    AVERAGE_WORD_LENGTH = TOTAL_CHARACTER_COUNT/Word_Count
    
    #Creating the array of output variables to be populated in Output DS
    Output_Array = [[POSITIVE_SCORE, NEGATIVE_SCORE, POLARITY_SCORE,
                    SUBJECTIVITY_SCORE, AVERAGE_SENTENCE_LENGTH, 
                    PERCENTAGE_OF_COMPLEX_WORDS, FOG_INDEX, 
                    AVERAGE_SENTENCE_LENGTH, COMPLEX_WORD_COUNT, 
                    CLEAN_WORD_COUNT, SYLLABLE_COUNT_PER_WORD,
                    PERSONAL_PRONOUN_COUNT, AVERAGE_WORD_LENGTH]]
    
    #Creating a data-frame of output array of that particular file
    Output_Data_Frame = pd.DataFrame(Output_Array)
    
    #Writing the output datastructure on to the Output Data structure File
    with pd.ExcelWriter('Output Data Structure.xlsx', engine='openpyxl', 
                        mode='a') as writer:
        book = load_workbook('Output Data Structure.xlsx')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        Output_Data_Frame.to_excel(writer, sheet_name='Sheet1', header=None, 
                                   index=False, startcol=2,startrow=index+1)
        
        writer.save()
        File.close()
        print(f"URL_{index}\tOutput generated for URL_{simplified_URL}\n")
        Log_File.log_event(f"URL_{index}\n")
        Log_File.log_event(f"Output generated for URL_{simplified_URL}\n")
        Log_File.log_event(Text_File_Generator.Separator)

Log_File.close()